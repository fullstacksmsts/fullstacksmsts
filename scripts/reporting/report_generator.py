#!/usr/bin/env python3
"""
SEO Report Generator for fullstacksmsts.co.uk
This script generates SEO reports based on analytics data and templates.
"""

import os
import logging
import argparse
import json
import pandas as pd
import numpy as np
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("Report_Generator")

# Load environment variables
load_dotenv()

class ReportGenerator:
    """Class to handle SEO report generation."""
    
    def __init__(self, start_date, end_date, report_type="monthly", output_file=None):
        """Initialize the report generator."""
        self.start_date = start_date
        self.end_date = end_date
        self.report_type = report_type
        self.output_file = output_file or f"reports/{report_type}_seo_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # Load configuration
        self.config = self._load_config()
        
        # Look for template in templates/reporting directory first, then in current directory
        template_paths = [
            f"templates/reporting/{report_type}_reporting_template.xlsx",
            f"templates/reporting/monthly_reporting_template.xlsx",  # Fallback to monthly template
            "monthly_reporting_template.xlsx"  # Original location for backward compatibility
        ]
        
        self.template_file = None
        for path in template_paths:
            if os.path.exists(path):
                self.template_file = path
                break
        
        if not self.template_file:
            logger.warning(f"Template file not found in any of the expected locations: {template_paths}")
            # Create a basic template if none exists
            self.template_file = "templates/reporting/monthly_reporting_template.xlsx"
            self._create_basic_template()
        
        self.data = {}
        
        # Create output directory if it doesn't exist
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
    
    def _load_config(self):
        """Load configuration from config.json file."""
        try:
            with open("config.json", "r") as f:
                config = json.load(f)
            return config
        except Exception as e:
            logger.error(f"Error loading configuration: {str(e)}")
            return {}
    
    def _create_basic_template(self):
        """Create a basic template if none exists."""
        try:
            # Create directory if it doesn't exist
            os.makedirs(os.path.dirname(self.template_file), exist_ok=True)
            
            # Create a new workbook
            wb = Workbook()
            
            # Add a sheet
            ws = wb.active
            ws.title = "SEO Report"
            
            # Add headers
            headers = ["Date", "Sessions", "Users", "New Users", "Bounce Rate", "Pages/Session", "Avg. Session Duration"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Save the workbook
            wb.save(self.template_file)
            logger.info(f"Created basic template at {self.template_file}")
        except Exception as e:
            logger.error(f"Error creating basic template: {str(e)}")
    
    def _load_analytics_data(self):
        """Load analytics data from CSV files."""
        logger.info("Loading analytics data...")
        
        # In a real implementation, this would load data from Google Analytics API
        # For this example, we'll generate sample data
        
        # Generate date range
        date_range = pd.date_range(start=self.start_date, end=self.end_date)
        
        # Generate organic traffic data
        organic_traffic = pd.DataFrame({
            'date': date_range,
            'sessions': np.random.randint(100, 500, size=len(date_range)),
            'users': np.random.randint(80, 400, size=len(date_range)),
            'new_users': np.random.randint(50, 200, size=len(date_range)),
            'bounce_rate': np.random.uniform(0.4, 0.7, size=len(date_range)),
            'pages_per_session': np.random.uniform(1.5, 4.0, size=len(date_range)),
            'avg_session_duration': np.random.uniform(60, 300, size=len(date_range))
        })
        
        # Store data in dictionary
        self.data = {
            'organic_traffic': organic_traffic
        }
        
        logger.info("Analytics data loaded successfully.")
        return self.data
    
    def upload_to_google_sheets(self, file_path):
        """Upload the report to Google Sheets."""
        try:
            # Import Google API libraries
            from googleapiclient.discovery import build
            from google.oauth2 import service_account
            
            # Get Google Sheets configuration
            google_sheets_config = self.config.get("reporting", {}).get("google_sheets", {})
            
            if not google_sheets_config.get("enabled", False):
                logger.info("Google Sheets integration is disabled in config.")
                return False
            
            credentials_file = google_sheets_config.get("credentials_file")
            spreadsheet_id = google_sheets_config.get("spreadsheet_id")
            
            if not credentials_file or not spreadsheet_id:
                logger.warning("Google Sheets credentials or spreadsheet ID not configured")
                return False
            
            if not os.path.exists(credentials_file):
                logger.error(f"Google credentials file not found: {credentials_file}")
                return False
            
            # Authenticate with Google Sheets API
            credentials = service_account.Credentials.from_service_account_file(
                credentials_file, 
                scopes=['https://www.googleapis.com/auth/spreadsheets']
            )
            
            # Build the Sheets service
            sheets_service = build('sheets', 'v4', credentials=credentials)
            
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Convert dataframe to values list
            values = [df.columns.tolist()] + df.values.tolist()
            
            # Create a new sheet with timestamp
            sheet_name = f"Report {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            
            # Add a new sheet
            request = sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "requests": [
                        {
                            "addSheet": {
                                "properties": {
                                    "title": sheet_name
                                }
                            }
                        }
                    ]
                }
            ).execute()
            
            # Get the new sheet ID
            sheet_id = request['replies'][0]['addSheet']['properties']['sheetId']
            
            # Update the values
            sheets_service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f"{sheet_name}!A1",
                valueInputOption="RAW",
                body={"values": values}
            ).execute()
            
            # Share with specified users if any
            share_with = google_sheets_config.get("share_with", [])
            if share_with:
                # Import drive API
                drive_service = build('drive', 'v3', credentials=credentials)
                
                for email in share_with:
                    drive_service.permissions().create(
                        fileId=spreadsheet_id,
                        body={
                            'type': 'user',
                            'role': 'reader',
                            'emailAddress': email
                        },
                        sendNotificationEmail=False
                    ).execute()
                    logger.info(f"Shared report with {email}")
            
            logger.info(f"Report uploaded to Google Sheets: {sheet_name}")
            return True
        
        except ImportError:
            logger.error("Google API libraries not installed. Run: pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
            return False
        except Exception as e:
            logger.error(f"Error uploading to Google Sheets: {str(e)}")
            return False
    
    def run(self):
        """Run the report generation process."""
        logger.info(f"Starting {self.report_type} report generation...")
        
        # Load analytics data
        self._load_analytics_data()
        
        # Load template workbook
        try:
            wb = load_workbook(self.template_file)
            logger.info(f"Template loaded from {self.template_file}")
        except Exception as e:
            logger.error(f"Error loading template: {str(e)}")
            # Try to create a new workbook if loading fails
            wb = Workbook()
            ws = wb.active
            ws.title = "SEO Report"
            logger.info("Created new workbook as fallback")
        
        # Save workbook
        try:
            wb.save(self.output_file)
            logger.info(f"Report saved to {self.output_file}")
            
            # Upload to Google Sheets if enabled
            if self.config.get("reporting", {}).get("google_sheets", {}).get("enabled", False):
                self.upload_to_google_sheets(self.output_file)
        except Exception as e:
            logger.error(f"Error saving report: {str(e)}")
            return False
        
        logger.info(f"{self.report_type.capitalize()} report generation completed successfully.")
        return True

def main(start_date=None, end_date=None, report_type="monthly", output_file=None):
    """Main function to run report generation."""
    # Set default dates if not provided
    if not start_date:
        # Default to first day of current month
        today = datetime.now()
        start_date = datetime(today.year, today.month, 1).strftime('%Y-%m-%d')
    
    if not end_date:
        # Default to today
        end_date = datetime.now().strftime('%Y-%m-%d')
    
    # Create and run report generator
    generator = ReportGenerator(start_date, end_date, report_type, output_file)
    return generator.run()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="SEO Report Generator for fullstacksmsts.co.uk")
    parser.add_argument("--start-date", help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end-date", help="End date (YYYY-MM-DD)")
    parser.add_argument("--report-type", default="monthly", choices=["monthly", "quarterly", "annual"], help="Report type")
    parser.add_argument("--output", help="Output file path")
    
    args = parser.parse_args()
    main(args.start_date, args.end_date, args.report_type, args.output)
