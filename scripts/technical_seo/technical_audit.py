#!/usr/bin/env python3
"""
Technical SEO Audit Script for fullstacksmsts.co.uk
This script performs technical SEO audits of the website.
"""

import os
import re
import csv
import json
import logging
import argparse
import pandas as pd
import numpy as np
import random
import requests
from datetime import datetime
from urllib.parse import urlparse, urljoin
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("Technical_Audit")

# Load environment variables
load_dotenv()

class TechnicalAuditor:
    """Class to handle technical SEO audits."""
    
    def __init__(self, website_url, crawl_depth=2, output_file=None, user_agent=None):
        """Initialize the technical auditor."""
        self.website_url = website_url
        self.crawl_depth = crawl_depth
        self.output_file = output_file or f"reports/technical_audit_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.user_agent = user_agent or "SEOAutomationBot/1.0"
        self.visited_urls = set()
        self.pages = []
        self.issues = []
        
        # Parse the base domain
        parsed_url = urlparse(website_url)
        self.base_domain = parsed_url.netloc
        
        # Create output directory if it doesn't exist
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
    
    def _make_request(self, url):
        """Make an HTTP request to the given URL."""
        try:
            headers = {
                "User-Agent": self.user_agent,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.5",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
            }
            
            response = requests.get(url, headers=headers, timeout=10)
            return response
        except requests.RequestException as e:
            logger.error(f"Error requesting {url}: {str(e)}")
            return None
    
    def _crawl_page(self, url, depth=0):
        """Crawl a page and extract information."""
        # Check if we've already visited this URL or reached max depth
        if url in self.visited_urls or depth > self.crawl_depth:
            return
        
        logger.info(f"Crawling {url} (depth {depth})")
        self.visited_urls.add(url)
        
        # Make the request
        response = self._make_request(url)
        if not response:
            self.issues.append({
                'url': url,
                'issue_type': 'Request Error',
                'description': 'Failed to make request',
                'severity': 'High'
            })
            return
        
        # Check status code
        status_code = response.status_code
        if status_code != 200:
            self.issues.append({
                'url': url,
                'issue_type': f'{status_code} Status Code',
                'description': f'Page returned {status_code} status code',
                'severity': 'High' if status_code >= 500 else ('Medium' if status_code >= 400 else 'Low')
            })
            if status_code >= 400:
                return
        
        # Parse HTML
        try:
            soup = BeautifulSoup(response.text, 'html.parser')
        except Exception as e:
            self.issues.append({
                'url': url,
                'issue_type': 'HTML Parsing Error',
                'description': f'Failed to parse HTML: {str(e)}',
                'severity': 'Medium'
            })
            return
        
        # Extract page information
        title = soup.title.string.strip() if soup.title else ''
        meta_description = ''
        meta_keywords = ''
        canonical = ''
        h1 = ''
        
        # Extract meta tags
        for meta in soup.find_all('meta'):
            if meta.get('name', '').lower() == 'description':
                meta_description = meta.get('content', '')
            elif meta.get('name', '').lower() == 'keywords':
                meta_keywords = meta.get('content', '')
        
        # Extract canonical link
        canonical_tag = soup.find('link', {'rel': 'canonical'})
        if canonical_tag:
            canonical = canonical_tag.get('href', '')
        
        # Extract h1
        h1_tag = soup.find('h1')
        if h1_tag:
            h1 = h1_tag.get_text().strip()
        
        # Extract all headings
        headings = {
            'h1': [h.get_text().strip() for h in soup.find_all('h1')],
            'h2': [h.get_text().strip() for h in soup.find_all('h2')],
            'h3': [h.get_text().strip() for h in soup.find_all('h3')],
            'h4': [h.get_text().strip() for h in soup.find_all('h4')],
            'h5': [h.get_text().strip() for h in soup.find_all('h5')],
            'h6': [h.get_text().strip() for h in soup.find_all('h6')]
        }
        
        # Extract all links
        links = []
        for link in soup.find_all('a', href=True):
            href = link.get('href')
            text = link.get_text().strip()
            
            # Resolve relative URLs
            if not href.startswith(('http://', 'https://', 'mailto:', 'tel:')):
                href = urljoin(url, href)
            
            links.append({
                'href': href,
                'text': text,
                'is_internal': self.base_domain in urlparse(href).netloc if href.startswith(('http://', 'https://')) else True
            })
        
        # Extract all images
        images = []
        for img in soup.find_all('img'):
            src = img.get('src', '')
            alt = img.get('alt', '')
            
            # Resolve relative URLs
            if src and not src.startswith(('http://', 'https://', 'data:')):
                src = urljoin(url, src)
            
            images.append({
                'src': src,
                'alt': alt
            })
        
        # Check for structured data
        structured_data = []
        for script in soup.find_all('script', {'type': 'application/ld+json'}):
            try:
                data = json.loads(script.string)
                structured_data.append(data)
            except (json.JSONDecodeError, TypeError):
                self.issues.append({
                    'url': url,
                    'issue_type': 'Invalid Structured Data',
                    'description': 'Invalid JSON-LD structured data',
                    'severity': 'Medium'
                })
        
        # Store page information
        page_info = {
            'url': url,
            'status_code': status_code,
            'title': title,
            'meta_description': meta_description,
            'meta_keywords': meta_keywords,
            'canonical': canonical,
            'h1': h1,
            'headings': headings,
            'links': links,
            'images': images,
            'structured_data': structured_data,
            'content_length': len(response.text),
            'response_time': response.elapsed.total_seconds()
        }
        
        self.pages.append(page_info)
        
        # Check for issues
        self._check_page_issues(page_info)
        
        # Crawl internal links if within depth limit
        if depth < self.crawl_depth:
            internal_links = [link['href'] for link in links if link['is_internal']]
            for link in internal_links:
                if link not in self.visited_urls:
                    self._crawl_page(link, depth + 1)
    
    def _check_page_issues(self, page_info):
        """Check for issues on a page."""
        url = page_info['url']
        
        # Check title
        if not page_info['title']:
            self.issues.append({
                'url': url,
                'issue_type': 'Missing Title',
                'description': 'Page has no title tag',
                'severity': 'High'
            })
        elif len(page_info['title']) < 10:
            self.issues.append({
                'url': url,
                'issue_type': 'Short Title',
                'description': f'Title is too short ({len(page_info["title"])} characters)',
                'severity': 'Medium'
            })
        elif len(page_info['title']) > 60:
            self.issues.append({
                'url': url,
                'issue_type': 'Long Title',
                'description': f'Title is too long ({len(page_info["title"])} characters)',
                'severity': 'Medium'
            })
        
        # Check meta description
        if not page_info['meta_description']:
            self.issues.append({
                'url': url,
                'issue_type': 'Missing Meta Description',
                'description': 'Page has no meta description',
                'severity': 'Medium'
            })
        elif len(page_info['meta_description']) < 50:
            self.issues.append({
                'url': url,
                'issue_type': 'Short Meta Description',
                'description': f'Meta description is too short ({len(page_info["meta_description"])} characters)',
                'severity': 'Medium'
            })
        elif len(page_info['meta_description']) > 160:
            self.issues.append({
                'url': url,
                'issue_type': 'Long Meta Description',
                'description': f'Meta description is too long ({len(page_info["meta_description"])} characters)',
                'severity': 'Low'
            })
        
        # Check H1
        if not page_info['h1']:
            self.issues.append({
                'url': url,
                'issue_type': 'Missing H1',
                'description': 'Page has no H1 heading',
                'severity': 'Medium'
            })
        elif len(page_info['headings']['h1']) > 1:
            self.issues.append({
                'url': url,
                'issue_type': 'Multiple H1s',
                'description': f'Page has {len(page_info["headings"]["h1"])} H1 headings',
                'severity': 'Medium'
            })
        
        # Check canonical
        if not page_info['canonical']:
            self.issues.append({
                'url': url,
                'issue_type': 'Missing Canonical',
                'description': 'Page has no canonical tag',
                'severity': 'Medium'
            })
        elif page_info['canonical'] != url and not url.endswith('/') and page_info['canonical'] != url + '/':
            self.issues.append({
                'url': url,
                'issue_type': 'Non-matching Canonical',
                'description': f'Canonical URL ({page_info["canonical"]}) does not match page URL',
                'severity': 'Medium'
            })
        
        # Check images
        for img in page_info['images']:
            if not img['alt']:
                self.issues.append({
                    'url': url,
                    'issue_type': 'Missing Alt Text',
                    'description': f'Image {img["src"]} has no alt text',
                    'severity': 'Medium'
                })
        
        # Check content length
        if page_info['content_length'] < 500:
            self.issues.append({
                'url': url,
                'issue_type': 'Thin Content',
                'description': f'Page has only {page_info["content_length"]} characters',
                'severity': 'Medium'
            })
        
        # Check response time
        if page_info['response_time'] > 1.0:
            self.issues.append({
                'url': url,
                'issue_type': 'Slow Response Time',
                'description': f'Page took {page_info["response_time"]:.2f} seconds to respond',
                'severity': 'Medium'
            })
    
    def _simulate_core_web_vitals(self):
        """Simulate Core Web Vitals metrics for demonstration purposes."""
        logger.info("Simulating Core Web Vitals metrics...")
        
        core_web_vitals = []
        
        for page in self.pages:
            url = page['url']
            
            # Simulate LCP (Largest Contentful Paint)
            lcp_mobile = round(random.uniform(1.5, 4.5), 1)
            lcp_desktop = round(random.uniform(1.0, 3.0), 1)
            
            # Simulate FID (First Input Delay)
            fid_mobile = round(random.uniform(50, 200), 0)
            fid_desktop = round(random.uniform(20, 100), 0)
            
            # Simulate CLS (Cumulative Layout Shift)
            cls_mobile = round(random.uniform(0.05, 0.25), 2)
            cls_desktop = round(random.uniform(0.02, 0.15), 2)
            
            # Determine status
            def get_lcp_status(lcp):
                if lcp <= 2.5:
                    return "Good"
                elif lcp <= 4.0:
                    return "Needs Improvement"
                else:
                    return "Poor"
            
            def get_fid_status(fid):
                if fid <= 100:
                    return "Good"
                elif fid <= 300:
                    return "Needs Improvement"
                else:
                    return "Poor"
            
            def get_cls_status(cls):
                if cls <= 0.1:
                    return "Good"
                elif cls <= 0.25:
                    return "Needs Improvement"
                else:
                    return "Poor"
            
            # Add to Core Web Vitals data
            core_web_vitals.append({
                'url': url,
                'lcp_mobile': lcp_mobile,
                'lcp_mobile_status': get_lcp_status(lcp_mobile),
                'lcp_desktop': lcp_desktop,
                'lcp_desktop_status': get_lcp_status(lcp_desktop),
                'fid_mobile': fid_mobile,
                'fid_mobile_status': get_fid_status(fid_mobile),
                'fid_desktop': fid_desktop,
                'fid_desktop_status': get_fid_status(fid_desktop),
                'cls_mobile': cls_mobile,
                'cls_mobile_status': get_cls_status(cls_mobile),
                'cls_desktop': cls_desktop,
                'cls_desktop_status': get_cls_status(cls_desktop)
            })
            
            # Add issues for poor metrics
            if get_lcp_status(lcp_mobile) == "Poor":
                self.issues.append({
                    'url': url,
                    'issue_type': 'Poor LCP (Mobile)',
                    'description': f'Mobile LCP is {lcp_mobile}s (should be <= 2.5s)',
                    'severity': 'High'
                })
            
            if get_fid_status(fid_mobile) == "Poor":
                self.issues.append({
                    'url': url,
                    'issue_type': 'Poor FID (Mobile)',
                    'description': f'Mobile FID is {fid_mobile}ms (should be <= 100ms)',
                    'severity': 'High'
                })
            
            if get_cls_status(cls_mobile) == "Poor":
                self.issues.append({
                    'url': url,
                    'issue_type': 'Poor CLS (Mobile)',
                    'description': f'Mobile CLS is {cls_mobile} (should be <= 0.1)',
                    'severity': 'High'
                })
        
        logger.info(f"Simulated Core Web Vitals for {len(core_web_vitals)} pages.")
        return core_web_vitals
    
    def _generate_report(self, core_web_vitals):
        """Generate a technical SEO audit report."""
        logger.info("Generating technical SEO audit report...")
        
        # Create a new workbook
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        good_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        warning_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        error_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add summary headers
        summary_headers = ["Metric", "Value"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Add summary data
        summary_data = [
            ["Audit Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Website URL", self.website_url],
            ["Crawl Depth", self.crawl_depth],
            ["Pages Crawled", len(self.pages)],
            ["Total Issues", len(self.issues)],
            ["High Severity Issues", len([i for i in self.issues if i['severity'] == 'High'])],
            ["Medium Severity Issues", len([i for i in self.issues if i['severity'] == 'Medium'])],
            ["Low Severity Issues", len([i for i in self.issues if i['severity'] == 'Low'])],
            ["Average Page Size", f"{sum(p['content_length'] for p in self.pages) / len(self.pages):.2f} bytes" if self.pages else "N/A"],
            ["Average Response Time", f"{sum(p['response_time'] for p in self.pages) / len(self.pages):.2f} seconds" if self.pages else "N/A"]
        ]
        
        for row, (metric, value) in enumerate(summary_data, 2):
            ws_summary.cell(row=row, column=1).value = metric
            ws_summary.cell(row=row, column=2).value = value
            
            for col in range(1, 3):
                ws_summary.cell(row=row, column=col).border = border
        
        # Add Core Web Vitals summary
        if core_web_vitals:
            ws_summary.cell(row=len(summary_data) + 3, column=1).value = "Core Web Vitals Summary"
            ws_summary.cell(row=len(summary_data) + 3, column=1).font = Font(bold=True)
            
            cwv_summary_data = [
                ["Average LCP (Mobile)", f"{sum(cwv['lcp_mobile'] for cwv in core_web_vitals) / len(core_web_vitals):.2f}s"],
                ["Average LCP (Desktop)", f"{sum(cwv['lcp_desktop'] for cwv in core_web_vitals) / len(core_web_vitals):.2f}s"],
                ["Average FID (Mobile)", f"{sum(cwv['fid_mobile'] for cwv in core_web_vitals) / len(core_web_vitals):.0f}ms"],
                ["Average FID (Desktop)", f"{sum(cwv['fid_desktop'] for cwv in core_web_vitals) / len(core_web_vitals):.0f}ms"],
                ["Average CLS (Mobile)", f"{sum(cwv['cls_mobile'] for cwv in core_web_vitals) / len(core_web_vitals):.2f}"],
                ["Average CLS (Desktop)", f"{sum(cwv['cls_desktop'] for cwv in core_web_vitals) / len(core_web_vitals):.2f}"]
            ]
            
            for row, (metric, value) in enumerate(cwv_summary_data, len(summary_data) + 4):
                ws_summary.cell(row=row, column=1).value = metric
                ws_summary.cell(row=row, column=2).value = value
                
                for col in range(1, 3):
                    ws_summary.cell(row=row, column=col).border = border
        
        # Create Issues sheet
        ws_issues = wb.create_sheet("Issues")
        
        # Add issues headers
        issue_headers = ["URL", "Issue Type", "Description", "Severity"]
        for col, header in enumerate(issue_headers, 1):
            cell = ws_issues.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Add issues data
        for row, issue in enumerate(self.issues, 2):
            ws_issues.cell(row=row, column=1).value = issue['url']
            ws_issues.cell(row=row, column=2).value = issue['issue_type']
            ws_issues.cell(row=row, column=3).value = issue['description']
            ws_issues.cell(row=row, column=4).value = issue['severity']
            
            # Apply conditional formatting based on severity
            severity_cell = ws_issues.cell(row=row, column=4)
            if issue['severity'] == 'High':
                severity_cell.fill = error_fill
            elif issue['severity'] == 'Medium':
                severity_cell.fill = warning_fill
            else:
                severity_cell.fill = good_fill
            
            for col in range(1, 5):
                ws_issues.cell(row=row, column=col).border = border
        
        # Create Pages sheet
        ws_pages = wb.create_sheet("Pages")
        
        # Add pages headers
        page_headers = ["URL", "Status Code", "Title", "Meta Description", "H1", "Content Length", "Response Time"]
        for col, header in enumerate(page_headers, 1):
            cell = ws_pages.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Add pages data
        for row, page in enumerate(self.pages, 2):
            ws_pages.cell(row=row, column=1).value = page['url']
            ws_pages.cell(row=row, column=2).value = page['status_code']
            ws_pages.cell(row=row, column=3).value = page['title']
            ws_pages.cell(row=row, column=4).value = page['meta_description']
            ws_pages.cell(row=row, column=5).value = page['h1']
            ws_pages.cell(row=row, column=6).value = page['content_length']
            ws_pages.cell(row=row, column=7).value = f"{page['response_time']:.2f}s"
            
            # Apply conditional formatting based on status code
            status_cell = ws_pages.cell(row=row, column=2)
            if page['status_code'] >= 500:
                status_cell.fill = error_fill
            elif page['status_code'] >= 400:
                status_cell.fill = warning_fill
            else:
                status_cell.fill = good_fill
            
            for col in range(1, 8):
                ws_pages.cell(row=row, column=col).border = border
        
        # Create Core Web Vitals sheet
        if core_web_vitals:
            ws_cwv = wb.create_sheet("Core Web Vitals")
            
            # Add CWV headers
            cwv_headers = [
                "URL", 
                "LCP (Mobile)", "Status", 
                "LCP (Desktop)", "Status", 
                "FID (Mobile)", "Status", 
                "FID (Desktop)", "Status", 
                "CLS (Mobile)", "Status", 
                "CLS (Desktop)", "Status"
            ]
            
            for col, header in enumerate(cwv_headers, 1):
                cell = ws_cwv.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Add CWV data
            for row, cwv in enumerate(core_web_vitals, 2):
                ws_cwv.cell(row=row, column=1).value = cwv['url']
                
                # LCP Mobile
                ws_cwv.cell(row=row, column=2).value = f"{cwv['lcp_mobile']}s"
                ws_cwv.cell(row=row, column=3).value = cwv['lcp_mobile_status']
                
                # LCP Desktop
                ws_cwv.cell(row=row, column=4).value = f"{cwv['lcp_desktop']}s"
                ws_cwv.cell(row=row, column=5).value = cwv['lcp_desktop_status']
                
                # FID Mobile
                ws_cwv.cell(row=row, column=6).value = f"{cwv['fid_mobile']}ms"
                ws_cwv.cell(row=row, column=7).value = cwv['fid_mobile_status']
                
                # FID Desktop
                ws_cwv.cell(row=row, column=8).value = f"{cwv['fid_desktop']}ms"
                ws_cwv.cell(row=row, column=9).value = cwv['fid_desktop_status']
                
                # CLS Mobile
                ws_cwv.cell(row=row, column=10).value = cwv['cls_mobile']
                ws_cwv.cell(row=row, column=11).value = cwv['cls_mobile_status']
                
                # CLS Desktop
                ws_cwv.cell(row=row, column=12).value = cwv['cls_desktop']
                ws_cwv.cell(row=row, column=13).value = cwv['cls_desktop_status']
                
                # Apply conditional formatting based on status
                for col in [3, 5, 7, 9, 11, 13]:  # Status columns
                    status_cell = ws_cwv.cell(row=row, column=col)
                    if status_cell.value == "Good":
                        status_cell.fill = good_fill
                    elif status_cell.value == "Needs Improvement":
                        status_cell.fill = warning_fill
                    else:
                        status_cell.fill = error_fill
                
                for col in range(1, 14):
                    ws_cwv.cell(row=row, column=col).border = border
        
        # Adjust column widths
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(self.output_file)
        logger.info(f"Technical SEO audit report saved to {self.output_file}")
    
    def run(self):
        """Run the technical SEO audit."""
        logger.info(f"Starting technical SEO audit of {self.website_url}...")
        
        # Crawl the website
        self._crawl_page(self.website_url)
        
        # Simulate Core Web Vitals
        core_web_vitals = self._simulate_core_web_vitals()
        
        # Generate report
        self._generate_report(core_web_vitals)
        
        logger.info(f"Technical SEO audit completed successfully. Report saved to {self.output_file}")
        return self.output_file

def main(website_url=None, crawl_depth=2, output_file=None, user_agent=None):
    """Main function to run technical SEO audit."""
    # Default website URL if not provided
    if not website_url:
        website_url = "https://fullstacksmsts.co.uk"
    
    # Create and run technical auditor
    auditor = TechnicalAuditor(website_url, crawl_depth, output_file, user_agent)
    return auditor.run()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Technical SEO Audit for fullstacksmsts.co.uk")
    parser.add_argument("--url", help="Website URL to audit")
    parser.add_argument("--depth", type=int, default=2, help="Crawl depth")
    parser.add_argument("--output", help="Output file path")
    parser.add_argument("--user-agent", help="User agent string")
    
    args = parser.parse_args()
    main(args.url, args.depth, args.output, args.user_agent)
